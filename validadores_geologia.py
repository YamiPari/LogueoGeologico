import streamlit as st
import pandas as pd
import csv
import chardet
import os
import io
import openpyxl
from openpyxl.styles import PatternFill, Font
import plotly.express as px


def leer_csv(archivo):
    try:
        if archivo is None:
            st.error("Error: No se ha subido ningún archivo.")
            return None

        # Intentar primero con utf-8
        try:
            archivo.seek(0)
            df = pd.read_csv(archivo, encoding="utf-8", on_bad_lines="skip")
        except UnicodeDecodeError:
            archivo.seek(0)
            df = pd.read_csv(archivo, encoding="utf-16", on_bad_lines="skip")

        if df.empty:
            st.error(f"Error: El archivo {archivo.name} está vacío.")
            return None

        return df
    except Exception as e:
        st.error(f"Error al leer el archivo {archivo.name}: {e}")
        return None

def validar_geo(df, hole_number):
    if df is None:
        st.error("⚠️ Error: No se pudo cargar el archivo Geology correctamente.")
        return None  # Detener la función si el DataFrame está vacío

    df.columns = df.columns.str.strip().str.lower()  # Normalizar nombres de columnas
    
    # Filtrar por HOLE_NUMBER
    df_filtrado = df[df["hole_number"] == hole_number]

    if df_filtrado.empty:
        st.warning(f"No se encontraron datos para HOLE_NUMBER: {hole_number}")
        return None

    condiciones = {
        31: ["VD"], 3: ["D", "D1"], 37: ["VAND"], 2: ["VL"], 28: ["VM"], 
        6: ["SPP"], 7: ["SOP"], 9: ["SPB"], 10: ["SOB"], 25: ["SSL"], 
        5: ["SSM"], 34: ["BXMM"], 30: ["I"], 14: ["P"], 8: ["BXC"], 
        32: ["VRD"], 33: ["VRD"], 12: ["CO"], 13: ["Q"], 17: ["LOST"], 15: ["F"]
    }

    df_filtrado['validación_geo'] = df_filtrado.apply(
        lambda row: 'correcto' if row['clito'] in condiciones and row['unit'] in condiciones[row['clito']] 
        else 'incorrecto', axis=1
    )

    return df_filtrado

# Función para validar Sample y Standards

def validar_sample_standards(sample_df, standards_df, hole_number):
    try:
        sample_df.columns = sample_df.columns.str.strip().str.lower()
        standards_df.columns = standards_df.columns.str.strip().str.lower()

        # Filtrar por HOLE_NUMBER
        sample_filtered = sample_df[sample_df['hole_number'] == hole_number].copy()
        standards_filtered = standards_df[standards_df['hole_number'] == hole_number].copy()

        # DataFrame para Sample
        df_sample = sample_filtered[['hole_number', 'sample_number', 'depth_from', 'depth_to', 'assay_sample_type_code','parent_sample_number']].copy()
        df_sample['tipo_muestra'] = df_sample['assay_sample_type_code']
        df_sample['depth_range'] = df_sample['depth_to'] - df_sample['depth_from']
        df_sample['tramo_valido'] = df_sample['depth_range'].apply(lambda x: '✅ Correcto' if 0.5 <= x <= 1.5 else '⚠️ Observado')
        df_sample = df_sample.drop(columns=['assay_sample_type_code'])

        # DataFrame para Standards (sin validación de tramo)
        df_standards = standards_filtered[['hole_number', 'sample_number', 'assay_standard_code']].copy()
        df_standards['tipo_muestra'] = df_standards['assay_standard_code']
        df_standards['depth_from'] = None
        df_standards['depth_to'] = None
        df_standards['depth_range'] = None
        df_standards['tramo_valido'] = None
        df_standards = df_standards.drop(columns=['assay_standard_code'])

        # Unir ambos DataFrames
        resultado = pd.concat([df_sample, df_standards], ignore_index=True)
        resultado = resultado.drop_duplicates(subset=['hole_number', 'sample_number', 'tipo_muestra'])
        resultado = resultado.sort_values(by='sample_number', ascending=True)

        # Reordenar columnas para visualización
        columnas_finales = [
            'hole_number', 'sample_number', 'tipo_muestra', 'parent_sample_number',
            'depth_from', 'depth_to', 'depth_range', 'tramo_valido'
        ]
        resultado = resultado[columnas_finales]

        return resultado
    except Exception as e:
        st.error(f"Error en validar_sample_standards: {e}")
        return None
    
# Función para validar Alteration
def validar_alteration(alteration_df, hole_number):
    try:
        if alteration_df is None:
            st.error("⚠️ Error: No se pudo cargar el archivo Alteration correctamente.")
            return None

        alteration_df.columns = alteration_df.columns.str.strip().str.lower()

        required_columns = ['hole_number', 'intensity_1', 'intensity_2', 'intensity_3', 
                            'distribution_1', 'distribution_2', 'distribution_3']
        missing_columns = [col for col in required_columns if col not in alteration_df.columns]

        if missing_columns:
            st.error(f"El archivo ALTERATION tiene columnas faltantes: {missing_columns}")
            return None

        # 🔹 Filtrar el DataFrame por `hole_number` antes de validar, asegurando que sea una copia independiente
        alteration_filtrado = alteration_df[alteration_df['hole_number'] == hole_number].copy()

        if alteration_filtrado.empty:
            st.warning(f"No se encontraron datos para HOLE_NUMBER: {hole_number}")
            return None

        def validar_filas(row):
            resultados = []
            for i in range(1, 4):
                if row[f'intensity_{i}'] == 'FORT' and row[f'distribution_{i}'] != 'PERV':
                    resultados.append(f"Incorrecto en intensity_{i} y distribution_{i} (esperado PERV)")
                if row[f'intensity_{i}'] == 'MODE' and pd.notnull(row[f'distribution_{i}']):
                    resultados.append(f"Incorrecto en intensity_{i} y distribution_{i} (esperado vacío)")
                if row[f'intensity_{i}'] == 'FRCA' and row[f'distribution_{i}'] not in ['PUNT', 'VEIN']:
                    resultados.append(f"Incorrecto en intensity_{i} y distribution_{i} (esperado PUNT o VEIN)")

            return " | ".join(resultados) if resultados else "Correcto"

        # 🔹 Aplicar la validación **solo a las filas filtradas**
        alteration_filtrado.loc[:, 'validación'] = alteration_filtrado.apply(validar_filas, axis=1)

        # 🔹 Retornar SOLO el DataFrame filtrado
        return alteration_filtrado  
    except Exception as e:
        st.error(f"Error durante la validación en ALTERATION: {e}")
        return None

# Mapeo entre Unit (Geology) y Rock_Type_Code (Major)
correspondencias = {
    "D": "ANDS", "VAND": "ANDS", "D1": "DIOR", "VL": "DACT", "VM": "DACT", "VD": "DACT",
    "SPP": "MASS", "SOP": "MASS", "SPB": "MASS", "SOB": "MASS", "SSL": "MASS", "SSM": "SMSS",
    "BXMM": "FSTF", "I": "GRDR", "P": "PEGM", "BXC": "BRTC", "VRD": "RIDC", "CO": "SOLO",
    "Q": "VTQZ", "LOST": "XXXX", "F": "PNZO", "LOST": "YYYY"
}


# Función para validar intervalos
def validar_intervalos(sample_df, validation_df, tipo, hole_number):
    try:
        sample_df.columns = sample_df.columns.str.strip().str.lower()
        validation_df.columns = validation_df.columns.str.strip().str.lower()

        sample_filtered = sample_df[sample_df['hole_number'] == hole_number]
        validation_filtered = validation_df[validation_df['hole_number'] == hole_number]

        if sample_filtered.empty:
            st.error(f"No se encontraron datos para HOLE_NUMBER {hole_number} en Sample.")
            return None
        if validation_filtered.empty:
            st.error(f"No se encontraron datos para HOLE_NUMBER {hole_number} en {tipo}.")
            return None

        sample_depth_from = sample_filtered['depth_from'].unique()
        sample_depth_to = sample_filtered['depth_to'].unique()

        resultados = []
        for _, row in validation_filtered.iterrows():
            depth_from_correcto = row['depth_from'] in sample_depth_from
            depth_to_correcto = row['depth_to'] in sample_depth_to

            validacion = "Correcto" if depth_from_correcto and depth_to_correcto else "Incorrecto"

            resultados.append({
                'hole_number': row['hole_number'],
                'depth_from': row['depth_from'],
                'depth_to': row['depth_to'],
                'archivo': tipo,
                'validación': validacion
            })

        return pd.DataFrame(resultados)
    except Exception as e:
        st.error(f"Error durante la validación de intervalos en {tipo}: {e}")
        return None

# Función para validar Major vs Geology
def validar_major_geology(geology_df, major_df, hole_number):
    try:
        geology_df.columns = geology_df.columns.str.strip().str.lower()
        major_df.columns = major_df.columns.str.strip().str.lower()

        geology_filtered = geology_df[geology_df['hole_number'] == hole_number]
        major_filtered = major_df[major_df['hole_number'] == hole_number]

        if geology_filtered.empty:
            st.error(f"No se encontraron datos para HOLE_NUMBER {hole_number} en Geology.")
            return None
        if major_filtered.empty:
            st.error(f"No se encontraron datos para HOLE_NUMBER {hole_number} en Major.")
            return None

        resultados = []
        for _, major_row in major_filtered.iterrows():
            major_from, major_to, rock_type = major_row['depth_from'], major_row['depth_to'], major_row['rock_type_code']

            geology_segmentos = geology_filtered[
                (geology_filtered['depth_from'] >= major_from) & (geology_filtered['depth_to'] <= major_to)
            ]

            if geology_segmentos.empty:
                validacion = "Incorrecto (No contiene segmentos de Geology)"
            else:
                unidades_geology = geology_segmentos['unit'].unique()
                validacion = "Correcto" if all(correspondencias.get(unit, "") == rock_type for unit in unidades_geology) else "Incorrecto (Rock_Type no coincide con Units de Geology)"

            resultados.append({
                'hole_number': hole_number,
                'depth_from_major': major_from,
                'depth_to_major': major_to,
                'rock_type_major': rock_type,
                'validación': validacion
            })

        return pd.DataFrame(resultados)
    except Exception as e:
        st.error(f"Error durante la validación entre Geology y Major: {e}")
        return None

# Interfaz en Streamlit
st.title("Validación de Datos Geológicos")

hole_number = st.text_input("Ingrese el HOLE_NUMBER a buscar:", key="hole_number_input")

# Carga de archivos en formato TXT y conversión a DataFrame
geology_file = st.file_uploader("Cargar Geology (.csv)", type=[".csv"], key="geology_uploader")
sample_file = st.file_uploader("Cargar Sample (.csv)", type=["csv"], key="sample_uploader")
standards_file = st.file_uploader("Cargar Standards (.csv)", type=["csv"], key="standards_uploader")
alteration_file = st.file_uploader("Cargar Alteration (.csv)", type=["csv"], key="alteration_uploader")
mine_file = st.file_uploader("Cargar Mine (.csv)", type=["csv"], key="mine_uploader")
major_file = st.file_uploader("Cargar Major (.csv)", type=["csv"], key="major_uploader")

# Convertir archivos TXT a DataFrames
geology_df = leer_csv(geology_file) if geology_file else None
sample_df = leer_csv(sample_file) if sample_file else None
standards_df = leer_csv(standards_file) if standards_file else None
alteration_df = leer_csv(alteration_file) if alteration_file else None
mine_df = leer_csv(mine_file) if mine_file else None
major_df = leer_csv(major_file) if major_file else None


# Función para descargar archivos
def descargar_resultados(df, nombre_archivo):
    if df is not None and not df.empty:
        # Crear un buffer en memoria
        output = io.BytesIO()
        
        # Guardar el DataFrame en el buffer como archivo Excel
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)
        
        # Mover el cursor al inicio del archivo
        output.seek(0)

        # Botón de descarga
        st.download_button(
            label=f"Descargar {nombre_archivo}",
            data=output,
            file_name=nombre_archivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

def exportar_a_excel(df, filename):
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Resultados', index=False)

            workbook = writer.book
            sheet = workbook['Resultados']

            # Formato del encabezado
            encabezado_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            encabezado_font = Font(bold=True)

            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=col)
                cell.value = str(cell.value).upper()
                cell.fill = encabezado_fill
                cell.font = encabezado_font

            # Colores para tipo_muestra
            colores = {
                "PECLSTD006": "F7F99F",
                "PECLSTD007": "3785BF",
                "RG": "F0DEF2",
                "DP": "B5E6A2"
            }

            tipo_col = None
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=col).value == "TIPO_MUESTRA":
                    tipo_col = col
                    break

            if tipo_col:
                for row in range(2, sheet.max_row + 1):
                    value = sheet.cell(row=row, column=tipo_col).value
                    if value in colores:
                        sheet.cell(row=row, column=tipo_col).fill = PatternFill(start_color=colores[value], end_color=colores[value], fill_type="solid")

            workbook.save(filename)

    except Exception as e:
        st.error(f"Error al exportar a Excel: {e}")
        
# Botones de validación con tablas interactivas
if st.button("Validar Geology", key="validate_geology") and geology_file:
    geology_df = leer_csv(geology_file)
    resultados_geo = validar_geo(geology_df, hole_number)
    st.dataframe(resultados_geo)  # Tabla interactiva
    descargar_resultados(resultados_geo, "resultados_geology.csv")

if st.button("Validar Sample & Standards", key="validate_sample_standards") and sample_file and standards_file:
    sample_df = leer_csv(sample_file)
    standards_df = leer_csv(standards_file)
    resultados_sample_standards = validar_sample_standards(sample_df, standards_df, hole_number)

    if resultados_sample_standards is not None:
        st.dataframe(resultados_sample_standards)
        exportar_a_excel(resultados_sample_standards, "PECLD07.xlsx")

        with open("PECLD07.xlsx", "rb") as file:
            st.download_button(label="⬇️ Descargar Excel",
                               data=file,
                               file_name="PECLD07.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
if st.button("Validar Alteration", key="validate_alteration") and alteration_file:
    alteration_df = leer_csv(alteration_file)
    resultados_alteration = validar_alteration(alteration_df, hole_number)
    st.dataframe(resultados_alteration)
    descargar_resultados(resultados_alteration, "resultados_alteration.csv")

if st.button("Validar Intervals", key="validate_intervals") and sample_file:
    sample_df = leer_csv(sample_file)
    
    archivos = {
        "Geology": geology_file,
        "Major": major_file,
        "Alteration": alteration_file,
        "Mine": mine_file
    }
    
    resultados_totales = []
    
    for tipo, archivo in archivos.items():
        if archivo:
            validation_df = leer_csv(archivo)
            resultados = validar_intervalos(sample_df, validation_df, tipo, hole_number)
            if resultados is not None:
                resultados_totales.append(resultados)
    
    if resultados_totales:
        resultados_finales = pd.concat(resultados_totales, ignore_index=True)
        st.dataframe(resultados_finales)
        descargar_resultados(resultados_finales, "resultados_validacion.csv")

if st.button("Validar Major", key="validate_major") and geology_file and major_file:
    geology_df = leer_csv(geology_file)
    major_df = leer_csv(major_file)

    if major_df is None or major_df.empty:
        st.error("Error: El archivo Major está vacío o no se pudo cargar correctamente.")
        st.stop()  # Detiene la ejecución sin errores

    resultados_major = validar_major_geology(geology_df, major_df, hole_number)
    if resultados_major is not None:
        st.dataframe(resultados_major)
        descargar_resultados(resultados_major, "resultados_major.csv")



# Lista de estándares a analizar
estandares_relevantes = ["PECLSTD006", "DP", "RG", "PECLSTD007", "PECLBLK002"]

# Función para calcular el porcentaje de estándares respecto a las muestras OR
def calcular_porcentaje_standards(sample_df, standards_df, hole_number):
    if sample_df is None or standards_df is None:
        st.warning("Error: No se han cargado ambos archivos correctamente.")
        return None, None

    # Filtrar por HOLE_NUMBER
    sample_filtrado = sample_df[sample_df["hole_number"] == hole_number]
    standards_filtrado = standards_df[standards_df["hole_number"] == hole_number]

    # Filtrar muestras OR (excluyendo DP y RG)
    muestras_or = sample_filtrado[~sample_filtrado["assay_sample_type_code"].isin(["DP", "RG"])]
    total_muestras_or = len(muestras_or)

    # Filtrar muestras DP y RG (que cuentan como estándares)
    muestras_dp_rg = sample_filtrado[sample_filtrado["assay_sample_type_code"].isin(["DP", "RG"])]
    total_dp_rg = len(muestras_dp_rg)

    total_standards = len(standards_filtrado)

    if total_muestras_or == 0 and total_standards == 0 and total_dp_rg == 0:
        st.warning(f"No se encontraron datos para HOLE_NUMBER: {hole_number}")
        return None, None

    # Calcular porcentaje con la ecuación correcta
    porcentaje_standards = (total_standards + total_dp_rg) / (total_standards + total_dp_rg + total_muestras_or) if (total_standards + total_dp_rg + total_muestras_or) > 0 else 0

    # Crear DataFrame con los resultados
    resumen_df = pd.DataFrame({
        "HOLE_NUMBER": [hole_number],
        "Total Muestras OR": [total_muestras_or],
        "Total Standards Relevantes": [total_standards],
        "Total DP/RG (como estándares)": [total_dp_rg],
        "Porcentaje Standards (%)": [porcentaje_standards * 100]
    })

    return resumen_df, porcentaje_standards


# Botón para validar Sample & Standards y calcular el porcentaje
if st.button("Ingreso de Sample & Standards", key="validate_sample_standards2") and hole_number:
    resultados_sample_standards = validar_sample_standards(sample_df, standards_df, hole_number)
    st.subheader("Resultados de validación:")
    st.dataframe(resultados_sample_standards)

    # 🔥 Nuevo análisis de porcentaje de estándares
    resumen_df, porcentaje = calcular_porcentaje_standards(sample_df, standards_df, hole_number)
    if resumen_df is not None:
        st.subheader("Resultados del análisis de estándares")
        st.dataframe(resumen_df)  # Tabla interactiva

        # Gráfico de barras con datos filtrados
        fig = px.bar(
            resumen_df.melt(value_vars=["Total Muestras OR", "Total Standards Relevantes", "Total DP/RG (como estándares)"]),
            x="variable", y="value", text="value",
            title=f"Comparación entre Muestras OR, Standards y DP/RG para HOLE_NUMBER {hole_number}",
            labels={"variable": "Tipo", "value": "Cantidad"},
            color="variable"
        )
        st.plotly_chart(fig)
